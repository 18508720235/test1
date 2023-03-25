import time

from openpyxl import load_workbook
from fuzzywuzzy import fuzz

import os

# 获取当前目录
current_dir = os.getcwd()
# 读取当前目录下的所有文件名
files_in_current_dir = os.listdir(current_dir)

# 将文件名存储到列表中
file_list = []
for filename in files_in_current_dir:
    file_list.append(filename)
print((file_list))
print(file_list[18:34])
# 显示文件列表


for fileName in file_list[18:34]:

    # 打开Excel文件并读取指定的工作表
    workbook = load_workbook(filename=fileName)
    worksheet = workbook["Sheet1"]
    print(fileName,'开始--------------------------------------')
    # 定义一个列表来存储所有单元格的字符串
    all_strings = []

    # 读取某一列的每个单元格的字符串，并将它们添加到列表中
    for column in worksheet.iter_cols(min_col=4, max_col=4, values_only=True):
        for cell_value in column:
            if cell_value is not None:
                all_strings.extend(cell_value.split("; "))


    # 计算重复数据的数量
    duplicates = len(all_strings) - len(set(all_strings))
    # print(all_strings)

    # 去除相似的字符串
    unique_strings = []
    for string in all_strings:
        if len(string) < 4:
            continue
        is_unique = True
        for unique_string in unique_strings:
            ratio = fuzz.token_set_ratio(string, unique_string)

            if ratio > 70:
                is_unique = False
                # print('@@@重复的字符串：',string)
                # print(len(string))
                break
        if is_unique:
            unique_strings.append(string)
            # print('没有重复的：',string)
            # print(len(string))


    # 计算所有数据和非重复数据的数量，并输出相关信息
    all_values = len(all_strings)
    non_duplicates = len(unique_strings)

    print("共有 {} 个数据".format(all_values))
    print("其中有 {} 行数据重复".format(duplicates))
    print("共有 {} 行数据不重复".format(non_duplicates))
    print(fileName,'结束*********************************')
    time.sleep(1)
